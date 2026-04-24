"""
报告生成 Skill (Report Generator Skill)
AutoClaw Skill ID: report_generator

职责：
  接收经过匹配评分的岗位列表（scored_jobs），
  将数据写入 Excel (.xlsx) 对比表，
  应用专业的条件格式（绿/黄/红评分色块）、冻结标题行、自适应列宽，
  输出《实习申请对比表》到 output/ 目录。

跨应用能力:
  GLM (数据) → Microsoft Excel (.xlsx 文件)

触发场景（通常由 Pipeline 自动触发）：
  - 在 match_scorer 执行完毕后自动触发
  - 或当用户说"帮我生成对比表 / 导出 Excel"

GLM 调度方式：
  GLM 将 scored_jobs 列表和 output_filename 传入此 Skill。
"""

import os
import pandas as pd
from typing import List, Dict, Any, Optional
from skills import AutoClawSkill

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    def get_column_letter(n): return chr(64 + n)


class ReportGeneratorSkill(AutoClawSkill):
    """
    报告生成 Skill
    将评分岗位数据写入 Excel，应用条件格式，导出到 output/ 目录
    """

    SKILL_NAME = "report_generator"
    SKILL_DESCRIPTION = "将匹配评分后的岗位列表写入 Excel 对比表，应用条件色块格式，导出到 output/ 目录。"

    # Excel 列定义（字段映射 + 列宽）
    COLUMN_CONFIG = [
        {"field": "tier",          "header": "推荐等级",   "width": 18},
        {"field": "company",       "header": "公司名称",   "width": 22},
        {"field": "title",         "header": "岗位名称",   "width": 26},
        {"field": "match_score",   "header": "匹配得分",   "width": 10},
        {"field": "match_reason",  "header": "综合评估建议", "width": 48},
        {"field": "salary",        "header": "薪资范围",   "width": 14},
        {"field": "location",      "header": "工作地点",   "width": 14},
        {"field": "source",        "header": "来源平台",   "width": 12},
        {"field": "requirements",  "header": "JD 核心要求", "width": 55},
    ]

    def run(self, scored_jobs: List[Dict], output_path: Optional[str] = None,
            keyword: str = "", city: str = "", **kwargs) -> Dict[str, Any]:
        """
        执行 Excel 报告生成。

        :param scored_jobs: match_scorer 输出的评分岗位列表
        :param output_path: 完整输出路径（可选，自动生成）
        :param keyword: 搜索关键词（用于生成文件名）
        :param city: 搜索城市（用于生成文件名）
        :return: 生成的文件路径
        """
        self.validate_input({"scored_jobs": scored_jobs}, ["scored_jobs"])

        if not scored_jobs:
            return self._error("scored_jobs 列表为空，无法生成报告")

        # 确定输出路径
        output_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "output"
        )
        os.makedirs(output_dir, exist_ok=True)

        if output_path is None:
            label = f"{city}_{keyword}".strip("_") or "实习"
            output_path = os.path.join(output_dir, f"{label}_实习申请对比表.xlsx")

        self.logger.info(f"开始生成 Excel 报告 | 岗位数：{len(scored_jobs)} | 输出：{output_path}")

        self._write_excel(scored_jobs, output_path)

        formatted = OPENPYXL_AVAILABLE
        return self._success(
            data={"file_path": output_path, "job_count": len(scored_jobs), "formatted": formatted},
            message=f"Excel 对比表已生成：{output_path}（共 {len(scored_jobs)} 个岗位）"
                    + ("" if formatted else "（未格式化，需安装 openpyxl）")
        )

    def _write_excel(self, jobs: List[Dict], output_path: str):
        """生成并格式化 Excel 文件"""
        # ── 构建 DataFrame ──
        rows = []
        for job in jobs:
            row = {col["field"]: job.get(col["field"], "") for col in self.COLUMN_CONFIG}
            # 截断过长的 JD 描述
            if row.get("requirements"):
                row["requirements"] = str(row["requirements"])[:300]
            rows.append(row)

        df = pd.DataFrame(rows)
        df.columns = [col["header"] for col in self.COLUMN_CONFIG]

        df.to_excel(output_path, index=False)

        if not OPENPYXL_AVAILABLE:
            self.logger.warning("openpyxl 未安装，跳过格式化步骤")
            return

        # ── openpyxl 格式化 ──
        wb = load_workbook(output_path)
        ws = wb.active

        # 样式定义
        HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
        CELL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
        CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
        THIN = Side(style="thin", color="BFBFBF")
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        # 条件色定义
        COLOR_GREEN = ("C6EFCE", "006100")   # 高分
        COLOR_YELLOW = ("FFEB9C", "9C5700")  # 中分
        COLOR_RED = ("FFC7CE", "9C0006")     # 低分

        # 列宽设置
        for i, col_cfg in enumerate(self.COLUMN_CONFIG, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = col_cfg["width"]

        # 标题行格式
        ws.row_dimensions[1].height = 28
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = BORDER

        # 数据行格式
        score_col_idx = next(
            (i + 1 for i, c in enumerate(self.COLUMN_CONFIG) if c["field"] == "match_score"),
            4
        )
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            ws.row_dimensions[row_idx].height = 50
            for col_idx, cell in enumerate(row, start=1):
                cell.border = BORDER
                if col_idx == score_col_idx:
                    cell.alignment = CENTER_ALIGN
                    score = cell.value
                    if isinstance(score, (int, float)):
                        if score >= 75:
                            bg, fg = COLOR_GREEN
                        elif score >= 50:
                            bg, fg = COLOR_YELLOW
                        else:
                            bg, fg = COLOR_RED
                        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                        cell.font = Font(bold=True, color=fg, size=11)
                else:
                    cell.alignment = CELL_ALIGN

        # 冻结首行
        ws.freeze_panes = "A2"

        # 工作表标签名
        ws.title = "实习申请对比表"

        wb.save(output_path)
        self.logger.info(f"Excel 格式化完成，已保存: {output_path}")
